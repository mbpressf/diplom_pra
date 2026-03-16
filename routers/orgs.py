import csv
import hashlib
import io
import json
import os
import secrets
import statistics
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy.orm import Session

from auth_utils import get_current_user
from database import get_db
from models import Category, OrgReportRun, Organization, OrganizationMember, Transaction, User, UserOrgConsent
from schemas import (
    OrgKpiSnapshot,
    OrgReportRunOut,
    OrganizationCreate,
    OrganizationDashboardOut,
    OrganizationJoin,
    OrganizationMembershipOut,
    OrganizationReportGenerate,
)

router = APIRouter(prefix="/orgs", tags=["Organizations"])

VALID_ROLES = {"owner", "manager", "viewer"}
VALID_STATUSES = {"active", "invited", "disabled"}
ANON_SALT = os.getenv("ORG_ANON_SALT", "finpotok-org-anon-v1")


def _safe_round(value: float) -> float:
    return round(float(value or 0), 2)


def _generate_invite_code(db: Session, length: int = 10) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    for _ in range(20):
        candidate = "".join(secrets.choice(alphabet) for _ in range(length))
        exists = db.query(Organization).filter(Organization.invite_code == candidate).first()
        if not exists:
            return candidate
    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Could not generate invite code")


def _get_org_or_404(db: Session, org_id: int) -> Organization:
    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization not found")
    return org


def _require_membership(
    db: Session,
    org_id: int,
    user_id: int,
    allowed_roles: set[str] | None = None,
) -> OrganizationMember:
    membership = (
        db.query(OrganizationMember)
        .filter(
            OrganizationMember.organization_id == org_id,
            OrganizationMember.user_id == user_id,
            OrganizationMember.status == "active",
        )
        .first()
    )
    if not membership:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied for this organization")

    if allowed_roles and membership.role not in allowed_roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient role for this action")

    return membership


def _serialize_membership(org: Organization, membership: OrganizationMember) -> OrganizationMembershipOut:
    role = membership.role if membership.role in VALID_ROLES else "viewer"
    member_status = membership.status if membership.status in VALID_STATUSES else "active"
    return OrganizationMembershipOut(
        id=org.id,
        name=org.name,
        industry=org.industry or "",
        invite_code=org.invite_code,
        member_role=role,
        member_status=member_status,
        created_at=org.created_at,
    )


def _resolve_period(
    period_type: Literal["week", "month"],
    start_date: date | None = None,
    end_date: date | None = None,
    anchor_date: date | None = None,
) -> tuple[date, date, date, date]:
    if start_date and end_date:
        start = start_date
        end = end_date
    elif start_date or end_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Provide both start_date and end_date")
    else:
        anchor = anchor_date or date.today()
        days = 7 if period_type == "week" else 30
        start = anchor - timedelta(days=days - 1)
        end = anchor

    if start > end:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid period: start_date > end_date")

    duration_days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=duration_days - 1)
    return start, end, prev_start, prev_end


def _anon_user_id(org_id: int, user_id: int) -> str:
    payload = f"{ANON_SALT}:{org_id}:{user_id}".encode("utf-8")
    return hashlib.sha256(payload).hexdigest()[:12]


def _savings_rate(income: float, expense: float) -> float:
    if income > 0:
        return ((income - expense) / income) * 100
    if expense > 0:
        return -100.0
    return 0.0


def _is_high_risk(income: float, expense: float, savings_rate_pct: float) -> bool:
    if savings_rate_pct <= -10:
        return True

    if income > 0:
        return (expense / income) >= 1.1

    return expense > 0


def _risk_level(income: float, expense: float, savings_rate_pct: float) -> str:
    if _is_high_risk(income, expense, savings_rate_pct):
        return "high"
    if savings_rate_pct < 10:
        return "medium"
    return "low"


def _collect_org_dataset(db: Session, org_id: int, start: date, end: date) -> dict:
    memberships = (
        db.query(OrganizationMember)
        .filter(
            OrganizationMember.organization_id == org_id,
            OrganizationMember.status == "active",
        )
        .all()
    )
    member_ids = [item.user_id for item in memberships]
    if not member_ids:
        return {
            "user_rows": [],
            "categories": [],
            "trend": [],
            "total_expense_rub": 0.0,
            "eligible_count": 0,
        }

    consented_ids = {
        item.user_id
        for item in db.query(UserOrgConsent)
        .filter(
            UserOrgConsent.organization_id == org_id,
            UserOrgConsent.user_id.in_(member_ids),
            UserOrgConsent.is_active.is_(True),
        )
        .all()
    }

    eligible_ids = [uid for uid in member_ids if uid in consented_ids]
    if not eligible_ids:
        return {
            "user_rows": [],
            "categories": [],
            "trend": [],
            "total_expense_rub": 0.0,
            "eligible_count": 0,
        }

    tx_rows = (
        db.query(Transaction, Category.name)
        .outerjoin(Category, Transaction.category_id == Category.id)
        .filter(
            Transaction.user_id.in_(eligible_ids),
            Transaction.date >= start,
            Transaction.date <= end,
        )
        .all()
    )

    user_stats = {
        uid: {
            "income": 0.0,
            "expense": 0.0,
            "tx_count": 0,
            "expense_by_category": defaultdict(float),
        }
        for uid in eligible_ids
    }
    category_expense = defaultdict(float)
    trend_map = defaultdict(lambda: {"income": 0.0, "expense": 0.0})

    for tx, category_name in tx_rows:
        stat = user_stats[tx.user_id]
        amount = float(tx.amount or 0)
        tx_date = tx.date
        stat["tx_count"] += 1

        if tx.type == "income":
            stat["income"] += amount
            trend_map[tx_date]["income"] += amount
        else:
            stat["expense"] += amount
            trend_map[tx_date]["expense"] += amount
            category = category_name or "Без категории"
            stat["expense_by_category"][category] += amount
            category_expense[category] += amount

    user_rows = []
    for user_id, stat in user_stats.items():
        if stat["tx_count"] == 0:
            continue

        income = _safe_round(stat["income"])
        expense = _safe_round(stat["expense"])
        balance = _safe_round(income - expense)
        savings_rate_pct = _safe_round(_savings_rate(income, expense))
        top_expense_category = "-"
        if stat["expense_by_category"]:
            top_expense_category = max(stat["expense_by_category"].items(), key=lambda item: item[1])[0]

        user_rows.append(
            {
                "anon_user_id": _anon_user_id(org_id, user_id),
                "income_rub": income,
                "expense_rub": expense,
                "balance_rub": balance,
                "savings_rate_pct": savings_rate_pct,
                "risk_level": _risk_level(income, expense, savings_rate_pct),
                "top_expense_category": top_expense_category,
                "tx_count": stat["tx_count"],
            }
        )

    user_rows.sort(key=lambda row: row["anon_user_id"])

    total_expense = sum(category_expense.values())
    category_rows = []
    for category, amount in sorted(category_expense.items(), key=lambda item: item[1], reverse=True):
        share = (amount / total_expense * 100) if total_expense > 0 else 0
        category_rows.append(
            {
                "category": category,
                "amount_rub": _safe_round(amount),
                "share_pct": _safe_round(share),
            }
        )

    trend_rows = []
    cursor = start
    while cursor <= end:
        totals = trend_map[cursor]
        income = _safe_round(totals["income"])
        expense = _safe_round(totals["expense"])
        trend_rows.append(
            {
                "date": cursor.isoformat(),
                "income_rub": income,
                "expense_rub": expense,
                "balance_rub": _safe_round(income - expense),
            }
        )
        cursor += timedelta(days=1)

    return {
        "user_rows": user_rows,
        "categories": category_rows,
        "trend": trend_rows,
        "total_expense_rub": _safe_round(total_expense),
        "eligible_count": len(eligible_ids),
    }


def _median(values: list[float]) -> float:
    if not values:
        return 0.0
    return _safe_round(statistics.median(values))


def _build_kpi_snapshot(current: dict, previous: dict) -> dict:
    rows = current["user_rows"]
    prev_rows = previous["user_rows"]

    income_values = [item["income_rub"] for item in rows]
    expense_values = [item["expense_rub"] for item in rows]
    savings_values = [item["savings_rate_pct"] for item in rows]
    prev_savings_values = [item["savings_rate_pct"] for item in prev_rows]

    active_users_count = len(rows)
    overspend_count = sum(1 for item in rows if item["expense_rub"] > item["income_rub"])
    high_risk_count = sum(1 for item in rows if item["risk_level"] == "high")

    current_median_savings = _median(savings_values)
    previous_median_savings = _median(prev_savings_values)

    top5_expense = sum(item["amount_rub"] for item in current["categories"][:5])
    total_expense = current["total_expense_rub"]

    overspend_share = (overspend_count / active_users_count * 100) if active_users_count else 0
    high_risk_share = (high_risk_count / active_users_count * 100) if active_users_count else 0
    top5_share = (top5_expense / total_expense * 100) if total_expense > 0 else 0

    return {
        "active_users_count": active_users_count,
        "median_income_rub": _median(income_values),
        "median_expense_rub": _median(expense_values),
        "median_savings_rate_pct": current_median_savings,
        "overspend_share_pct": _safe_round(overspend_share),
        "high_risk_share_pct": _safe_round(high_risk_share),
        "top5_expense_categories_share_pct": _safe_round(top5_share),
        "savings_rate_delta_vs_prev_period_pct": _safe_round(current_median_savings - previous_median_savings),
    }


def _parse_kpi_snapshot(raw: str | None) -> dict:
    defaults = {
        "active_users_count": 0,
        "median_income_rub": 0.0,
        "median_expense_rub": 0.0,
        "median_savings_rate_pct": 0.0,
        "overspend_share_pct": 0.0,
        "high_risk_share_pct": 0.0,
        "top5_expense_categories_share_pct": 0.0,
        "savings_rate_delta_vs_prev_period_pct": 0.0,
    }
    try:
        decoded = json.loads(raw or "{}")
    except json.JSONDecodeError:
        decoded = {}

    for key in defaults:
        if key in decoded:
            defaults[key] = decoded[key]
    return defaults


def _serialize_report_run(run: OrgReportRun) -> OrgReportRunOut:
    snapshot = _parse_kpi_snapshot(run.kpi_snapshot)
    return OrgReportRunOut(
        id=run.id,
        organization_id=run.organization_id,
        period_type=run.period_type,
        period_start=run.period_start,
        period_end=run.period_end,
        generated_at=run.generated_at,
        users_csv_link=run.users_csv_link,
        report_xlsx_link=run.report_xlsx_link,
        report_pdf_link=run.report_pdf_link,
        kpi_snapshot=OrgKpiSnapshot(**snapshot),
    )


def _resolve_export_period(
    db: Session,
    org_id: int,
    period_type: Literal["week", "month"],
    start_date: date | None,
    end_date: date | None,
    report_run_id: int | None,
) -> tuple[Literal["week", "month"], date, date, date, date]:
    if report_run_id:
        report_run = (
            db.query(OrgReportRun)
            .filter(
                OrgReportRun.id == report_run_id,
                OrgReportRun.organization_id == org_id,
            )
            .first()
        )
        if not report_run:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report run not found")

        _, _, prev_start, prev_end = _resolve_period(
            report_run.period_type,
            start_date=report_run.period_start,
            end_date=report_run.period_end,
        )
        return report_run.period_type, report_run.period_start, report_run.period_end, prev_start, prev_end

    start, end, prev_start, prev_end = _resolve_period(period_type, start_date=start_date, end_date=end_date)
    return period_type, start, end, prev_start, prev_end


def _build_report_payload(
    db: Session,
    org_id: int,
    period_type: Literal["week", "month"],
    start_date: date,
    end_date: date,
    prev_start: date,
    prev_end: date,
) -> tuple[dict, dict, dict]:
    current = _collect_org_dataset(db, org_id, start_date, end_date)
    previous = _collect_org_dataset(db, org_id, prev_start, prev_end)
    kpi = _build_kpi_snapshot(current, previous)
    return current, previous, kpi


def _build_report_links(org_id: int, period_type: str, period_start: date, period_end: date) -> tuple[str, str, str]:
    query = f"period_type={period_type}&start_date={period_start.isoformat()}&end_date={period_end.isoformat()}"
    return (
        f"/orgs/{org_id}/exports/users.csv?{query}",
        f"/orgs/{org_id}/exports/report.xlsx?{query}",
        f"/orgs/{org_id}/exports/report.pdf?{query}",
    )


@router.post("", response_model=OrganizationMembershipOut, status_code=status.HTTP_201_CREATED)
def create_organization(
    payload: OrganizationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org = Organization(
        name=payload.name.strip(),
        industry=payload.industry.strip(),
        invite_code=_generate_invite_code(db),
    )
    db.add(org)
    db.flush()

    membership = OrganizationMember(
        organization_id=org.id,
        user_id=current_user.id,
        role="owner",
        status="active",
    )
    consent = UserOrgConsent(
        organization_id=org.id,
        user_id=current_user.id,
        is_active=True,
        consent_given_at=datetime.now(timezone.utc),
    )
    db.add(membership)
    db.add(consent)
    db.commit()
    db.refresh(org)
    db.refresh(membership)

    return _serialize_membership(org, membership)


@router.post("/join", response_model=OrganizationMembershipOut)
def join_organization(
    payload: OrganizationJoin,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not payload.consent:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Consent is required for anonymized analytics",
        )

    org = db.query(Organization).filter(Organization.invite_code == payload.invite_code).first()
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Organization with invite code not found")

    membership = (
        db.query(OrganizationMember)
        .filter(
            OrganizationMember.organization_id == org.id,
            OrganizationMember.user_id == current_user.id,
        )
        .first()
    )
    if membership:
        membership.status = "active"
        if membership.role not in VALID_ROLES:
            membership.role = "viewer"
    else:
        membership = OrganizationMember(
            organization_id=org.id,
            user_id=current_user.id,
            role="viewer",
            status="active",
        )
        db.add(membership)

    consent = (
        db.query(UserOrgConsent)
        .filter(
            UserOrgConsent.organization_id == org.id,
            UserOrgConsent.user_id == current_user.id,
        )
        .first()
    )
    if consent:
        consent.is_active = True
        consent.consent_given_at = datetime.now(timezone.utc)
    else:
        db.add(
            UserOrgConsent(
                organization_id=org.id,
                user_id=current_user.id,
                is_active=True,
                consent_given_at=datetime.now(timezone.utc),
            )
        )

    db.commit()
    db.refresh(org)
    db.refresh(membership)
    return _serialize_membership(org, membership)


@router.get("/me", response_model=list[OrganizationMembershipOut])
def list_my_organizations(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    memberships = (
        db.query(OrganizationMember)
        .join(Organization, Organization.id == OrganizationMember.organization_id)
        .filter(
            OrganizationMember.user_id == current_user.id,
            OrganizationMember.status == "active",
        )
        .order_by(Organization.created_at.desc())
        .all()
    )

    return [_serialize_membership(item.organization, item) for item in memberships]


@router.get("/{org_id}/dashboard", response_model=OrganizationDashboardOut)
def get_org_dashboard(
    org_id: int,
    period_type: Literal["week", "month"] = Query(default="month"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    org = _get_org_or_404(db, org_id)
    _require_membership(db, org_id, current_user.id)

    start, end, prev_start, prev_end = _resolve_period(period_type, start_date=start_date, end_date=end_date)
    current, previous, kpi = _build_report_payload(db, org_id, period_type, start, end, prev_start, prev_end)
    _ = (current, previous)  # keeps payload available for debugging if needed

    return OrganizationDashboardOut(
        organization_id=org.id,
        organization_name=org.name,
        period_type=period_type,
        period_start=start,
        period_end=end,
        **kpi,
    )


@router.post("/{org_id}/reports/generate", response_model=OrgReportRunOut)
def generate_org_report(
    org_id: int,
    payload: OrganizationReportGenerate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_org_or_404(db, org_id)
    _require_membership(db, org_id, current_user.id, allowed_roles={"owner", "manager"})

    start, end, prev_start, prev_end = _resolve_period(payload.period_type, anchor_date=payload.end_date)
    current, previous, kpi = _build_report_payload(db, org_id, payload.period_type, start, end, prev_start, prev_end)
    _ = (current, previous)

    users_csv_link, report_xlsx_link, report_pdf_link = _build_report_links(org_id, payload.period_type, start, end)
    report_run = OrgReportRun(
        organization_id=org_id,
        period_type=payload.period_type,
        period_start=start,
        period_end=end,
        users_csv_link=users_csv_link,
        report_xlsx_link=report_xlsx_link,
        report_pdf_link=report_pdf_link,
        kpi_snapshot=json.dumps(kpi, ensure_ascii=False),
    )
    db.add(report_run)
    db.commit()
    db.refresh(report_run)
    return _serialize_report_run(report_run)


@router.get("/{org_id}/reports", response_model=list[OrgReportRunOut])
def list_org_reports(
    org_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_org_or_404(db, org_id)
    _require_membership(db, org_id, current_user.id)

    report_runs = (
        db.query(OrgReportRun)
        .filter(OrgReportRun.organization_id == org_id)
        .order_by(OrgReportRun.generated_at.desc(), OrgReportRun.id.desc())
        .all()
    )
    return [_serialize_report_run(item) for item in report_runs]


@router.get("/{org_id}/exports/users.csv")
def export_org_users_csv(
    org_id: int,
    period_type: Literal["week", "month"] = Query(default="month"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    report_run_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_org_or_404(db, org_id)
    _require_membership(db, org_id, current_user.id)

    _, start, end, prev_start, prev_end = _resolve_export_period(
        db,
        org_id,
        period_type,
        start_date,
        end_date,
        report_run_id,
    )
    current, _, _ = _build_report_payload(db, org_id, period_type, start, end, prev_start, prev_end)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "period_start",
            "period_end",
            "anon_user_id",
            "income_rub",
            "expense_rub",
            "balance_rub",
            "savings_rate_pct",
            "risk_level",
            "top_expense_category",
            "tx_count",
        ]
    )

    for row in current["user_rows"]:
        writer.writerow(
            [
                start.isoformat(),
                end.isoformat(),
                row["anon_user_id"],
                row["income_rub"],
                row["expense_rub"],
                row["balance_rub"],
                row["savings_rate_pct"],
                row["risk_level"],
                row["top_expense_category"],
                row["tx_count"],
            ]
        )

    filename = f"org_{org_id}_users_{start.isoformat()}_{end.isoformat()}.csv"
    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{org_id}/exports/report.xlsx")
def export_org_report_xlsx(
    org_id: int,
    period_type: Literal["week", "month"] = Query(default="month"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    report_run_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_org_or_404(db, org_id)
    _require_membership(db, org_id, current_user.id)

    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="openpyxl is not installed") from exc

    _, start, end, prev_start, prev_end = _resolve_export_period(
        db,
        org_id,
        period_type,
        start_date,
        end_date,
        report_run_id,
    )
    current, _, kpi = _build_report_payload(db, org_id, period_type, start, end, prev_start, prev_end)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Организация", org_id])
    overview.append(["Период", f"{start.isoformat()} — {end.isoformat()}"])
    overview.append([])
    overview.append(["KPI", "Value"])
    overview.append(["active_users_count", kpi["active_users_count"]])
    overview.append(["median_income_rub", kpi["median_income_rub"]])
    overview.append(["median_expense_rub", kpi["median_expense_rub"]])
    overview.append(["median_savings_rate_pct", kpi["median_savings_rate_pct"]])
    overview.append(["overspend_share_pct", kpi["overspend_share_pct"]])
    overview.append(["high_risk_share_pct", kpi["high_risk_share_pct"]])
    overview.append(["top5_expense_categories_share_pct", kpi["top5_expense_categories_share_pct"]])
    overview.append(["savings_rate_delta_vs_prev_period_pct", kpi["savings_rate_delta_vs_prev_period_pct"]])

    users_sheet = workbook.create_sheet("Users_Anon")
    users_sheet.append(
        [
            "period_start",
            "period_end",
            "anon_user_id",
            "income_rub",
            "expense_rub",
            "balance_rub",
            "savings_rate_pct",
            "risk_level",
            "top_expense_category",
            "tx_count",
        ]
    )
    for row in current["user_rows"]:
        users_sheet.append(
            [
                start.isoformat(),
                end.isoformat(),
                row["anon_user_id"],
                row["income_rub"],
                row["expense_rub"],
                row["balance_rub"],
                row["savings_rate_pct"],
                row["risk_level"],
                row["top_expense_category"],
                row["tx_count"],
            ]
        )

    categories_sheet = workbook.create_sheet("Categories")
    categories_sheet.append(["category", "amount_rub", "share_pct"])
    for row in current["categories"]:
        categories_sheet.append([row["category"], row["amount_rub"], row["share_pct"]])

    trend_sheet = workbook.create_sheet("Trend")
    trend_sheet.append(["date", "income_rub", "expense_rub", "balance_rub"])
    for row in current["trend"]:
        trend_sheet.append([row["date"], row["income_rub"], row["expense_rub"], row["balance_rub"]])

    content = io.BytesIO()
    workbook.save(content)
    content.seek(0)

    filename = f"org_{org_id}_report_{start.isoformat()}_{end.isoformat()}.xlsx"
    return Response(
        content=content.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{org_id}/exports/report.pdf")
def export_org_report_pdf(
    org_id: int,
    period_type: Literal["week", "month"] = Query(default="month"),
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    report_run_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_org_or_404(db, org_id)
    _require_membership(db, org_id, current_user.id)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="reportlab is not installed") from exc

    _, start, end, prev_start, prev_end = _resolve_export_period(
        db,
        org_id,
        period_type,
        start_date,
        end_date,
        report_run_id,
    )
    current, _, kpi = _build_report_payload(db, org_id, period_type, start, end, prev_start, prev_end)

    pdf_stream = io.BytesIO()
    pdf = canvas.Canvas(pdf_stream, pagesize=A4)

    font_name = "Helvetica"
    for font_path in (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
    ):
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont("FinPotokSans", font_path))
            font_name = "FinPotokSans"
            break

    y = 805

    def draw_line(text: str, size: int = 11, gap: int = 16):
        nonlocal y
        if y <= 50:
            pdf.showPage()
            y = 805
        pdf.setFont(font_name, size)
        pdf.drawString(44, y, text)
        y -= gap

    draw_line("Итог периода", size=14, gap=22)
    draw_line(f"Организация: {org_id}")
    draw_line(f"Период: {start.isoformat()} — {end.isoformat()}")
    draw_line(f"Активных пользователей: {kpi['active_users_count']}")

    y -= 8
    draw_line("KPI", size=14, gap=22)
    draw_line(f"median_income_rub: {kpi['median_income_rub']}")
    draw_line(f"median_expense_rub: {kpi['median_expense_rub']}")
    draw_line(f"median_savings_rate_pct: {kpi['median_savings_rate_pct']}")
    draw_line(f"overspend_share_pct: {kpi['overspend_share_pct']}")
    draw_line(f"high_risk_share_pct: {kpi['high_risk_share_pct']}")
    draw_line(f"top5_expense_categories_share_pct: {kpi['top5_expense_categories_share_pct']}")
    draw_line(f"savings_rate_delta_vs_prev_period_pct: {kpi['savings_rate_delta_vs_prev_period_pct']}")

    y -= 8
    draw_line("Риски", size=14, gap=22)
    if kpi["high_risk_share_pct"] >= 30:
        draw_line("Высокая доля сотрудников в зоне финансового риска.")
    elif kpi["high_risk_share_pct"] >= 15:
        draw_line("Средняя доля сотрудников в зоне риска, нужен фокус на профилактике.")
    else:
        draw_line("Доля сотрудников в зоне риска контролируемая.")

    if kpi["overspend_share_pct"] >= 40:
        draw_line("Наблюдается существенный перерасход в текущем периоде.")
    else:
        draw_line("Перерасход находится в допустимом диапазоне.")

    y -= 8
    draw_line("Рекомендуемые действия", size=14, gap=22)
    draw_line("1. Внедрить еженедельный мониторинг KPI и оповещения.")
    draw_line("2. Провести targeted-программы финансовой грамотности для high-risk группы.")
    draw_line("3. Разобрать топ-5 категорий расходов и сформировать рекомендации по снижению трат.")

    if current["categories"]:
        y -= 8
        draw_line("Топ категории расходов", size=13, gap=20)
        for row in current["categories"][:5]:
            draw_line(f"- {row['category']}: {row['amount_rub']} RUB ({row['share_pct']}%)")

    pdf.save()
    pdf_stream.seek(0)

    filename = f"org_{org_id}_report_{start.isoformat()}_{end.isoformat()}.pdf"
    return Response(
        content=pdf_stream.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
